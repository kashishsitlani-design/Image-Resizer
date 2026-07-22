"""
Creative Editing for Atlas
Keeps the original image look by default.
No forced white background. No shadow. No cropping unless the user chooses Remove padding.

Naming rule (all sources behave the same way):
  - Uploaded file        -> keeps its original filename
  - Excel link           -> uses the value from the detected filename column, as-is
  - Pasted link          -> uses the name you give it ("name,url"), or the URL's own filename
  A numeric suffix is only added if two outputs would otherwise collide.

Requirements: streamlit, Pillow, pandas, requests, openpyxl, rembg, onnxruntime
"""
import gc
import io
import re
import time
import zipfile
from collections import Counter
from typing import Dict, Iterable, List, Optional, Tuple
from urllib.parse import urlparse, unquote

import pandas as pd
import requests
import streamlit as st
from PIL import Image, ImageChops, ImageOps, UnidentifiedImageError

try:
    from rembg import remove as rembg_remove, new_session
    REMBG_AVAILABLE = True
except Exception:
    REMBG_AVAILABLE = False

st.set_page_config(page_title="Creative Editing for Atlas", page_icon="🎨", layout="centered")

NAVY, LIGHT_NAVY, BORDER, SOFT_BG = "#0B2E59", "#123E73", "#D9E2EC", "#F5F8FC"
st.markdown(
    f"""
    <style>
    .stApp {{ background: white; }}
    h1, h2, h3 {{ color: {NAVY}; }}
    .stButton > button, .stDownloadButton > button {{
        background-color: {NAVY} !important; color: white !important;
        border: 1px solid {NAVY} !important; border-radius: 8px !important; font-weight: 600 !important;
    }}
    .stButton > button:hover, .stDownloadButton > button:hover {{
        background-color: {LIGHT_NAVY} !important; border-color: {LIGHT_NAVY} !important; color: white !important;
    }}
    div[data-testid="stMetric"] {{ background: {SOFT_BG}; border: 1px solid {BORDER}; border-radius: 10px; padding: 12px; }}
    div[data-testid="stFileUploader"] section {{ border-color: {BORDER}; background: {SOFT_BG}; }}
    .small-note {{ font-size: 0.86rem; color: #4B5563; }}
    </style>
    """,
    unsafe_allow_html=True,
)

st.title("Creative Editing for Atlas")
st.caption("Resize images safely • Exact marketplace sizes • No stretch • Excel/pasted links • ZIP download")

# Marketplace -> Marketplace+Region code, straight from PXM_Marketplace_Image_Naming_Convention.xlsx
# (the "Data" sheet, column D). This is the code that goes into the filename and into the
# Image Stack Group field, so it must exactly match what's in that reference file.
PXM_MARKETPLACE_CODES = {
    "Allegro PL": "ALG_PL",
    "Allegro One PL": "ALGO_PL",
    "Best Buy US": "BSTB_US",
    "Best Buy CA": "BSTB_CA",
    "Bol NL": "BOL_NL",
    "Bol BE": "BOL_BE",
    "eBay US": "EBY_US",
    "eBay DE": "EBY_DE",
    "eBay UK": "EBY_UK",
    "Kohl's US": "KHL_US",
    "Lowes US": "LWS_US",
    "Macy's US": "MCY_US",
    "MediaMarkt DE": "MM_DE",
    "Mercado Libre US": "MLBR_US",
    "Nordstrom US": "NRDM_US",
    "Octopia FR": "OCT_FR",
    "OTTO DE": "OTTO_DE",
    "Target US": "TGT_US",
    "Tesco UK": "TSC_UK",
    "Tik Tok US": "TTS_US",
    "Tik Tok UK": "TTS_UK",
    "Walmart US": "WM_US",
    "Walmart CA": "WM_CA",
    "Zalando DE": "ZL_DE",
    "NOON AE": "NOON_AE",
    "Shopify": "SPFY_US",
}
STACK_TEMPLATE_HEADERS = ["Import Type", "Product", "Image Stack Group", "Filename", "Image Stack Order"]

workflow = st.radio(
    "What do you want to do?",
    ["Resize / edit images", "Download, rename for PXM & build Image Stack template"],
    horizontal=True,
)

# name -> (recommended_w, recommended_h, min, max, aspect_ratio)
MARKETPLACE_PRESETS = {
    "Custom": (800, 800, "—", "—", "—"),
    "Allegro PL — 2200 x 2200": (2200, 2200, "500 x 500", "2560 x 2560", "1:1"),
    "Allegro One PL — 2200 x 2200": (2200, 2200, "1000 x 1000", "2560 x 2560", "1:1"),
    "Best Buy US — 2000 x 2000": (2000, 2000, "2000 x 2000", "—", "1:1"),
    "Best Buy CA — 2000 x 2000": (2000, 2000, "2000 x 2000", "—", "1:1"),
    "Bol NL — 2400 x 2400": (2400, 2400, "500 x 500", "6000 x 6000", "1:1"),
    "Bol BE — 2400 x 2400": (2400, 2400, "500 x 500", "6000 x 6000", "1:1"),
    "eBay US — 1600 x 1600": (1600, 1600, "500 x 500", "9000 x 9000", "1:1"),
    "eBay DE — 1600 x 1600": (1600, 1600, "500 x 500", "9000 x 9000", "1:1"),
    "eBay UK — 1600 x 1600": (1600, 1600, "500 x 500", "9000 x 9000", "1:1"),
    "Kohl's US — 1000 x 1000": (1000, 1000, "1000 x 1000", "—", "1:1"),
    "Lowes US — 1000 x 1000": (1000, 1000, "1000 x 1000", "—", "1:1"),
    "Macy's US — 1000 x 1000": (1000, 1000, "1000 x 1000", "—", "1:1"),
    "MediaMarkt DE — 1200 x 1200": (1200, 1200, "1000 x 1000", "—", "1:1"),
    "Mercado Libre US — 1600 x 1600": (1600, 1600, "500 x 500", "2500 x 2500", "1:1"),
    "Nordstrom US — 2600 x 4000": (2600, 4000, "1300 x 2000", "—", "2:3"),
    "Octopia FR — 500 x 500": (500, 500, "1000 x 1000", "2500 x 2500", "1:1"),
    "OTTO DE — 960 x 480": (960, 480, "—", "—", "2:1"),
    "Target US — 2400 x 2400": (2400, 2400, "1200 x 1200", "5000 x 5000", "1:1"),
    "Tesco UK — 2400 x 2400": (2400, 2400, "1000 x 1000", "—", "1:1"),
    "Tik Tok US — 1000 x 1000": (1000, 1000, "600 x 600", "3000 x 3000", "1:1"),
    "Tik Tok UK — 1000 x 1000": (1000, 1000, "600 x 600", "3000 x 3000", "1:1"),
    "Walmart US — 2200 x 2200": (2200, 2200, "1500 x 1500", "5000 x 5000", "1:1"),
    "Walmart CA — 2200 x 2200": (2200, 2200, "1500 x 1500", "5000 x 5000", "1:1"),
    "Zalando DE — 2000 x 2000": (2000, 2000, "800 x 1200", "5000 x 5000", "2:3"),
}
PRESETS = {name: (info[0], info[1]) for name, info in MARKETPLACE_PRESETS.items()}
EXT_MAP = {"PNG": "png", "JPEG": "jpg", "WEBP": "webp", "GIF": "gif", "BMP": "bmp", "TIFF": "tif"}
URL_KEYWORDS = ("image", "img", "link", "url", "photo", "picture", "media", "front", "back", "side", "lifestyle")
FILE_KEYWORDS = ("filename", "file name", "file", "sku", "productid", "product id", "item", "item code", "code", "name", "title")


# ---------- small shared helpers ----------

def clean_text(value) -> str:
    if pd.isna(value):
        return ""
    return str(value).strip()


def is_url(value: str) -> bool:
    parsed = urlparse(value.strip()) if value else None
    return bool(parsed and parsed.scheme in ("http", "https") and parsed.netloc)


def safe_filename(name: str, fallback: str = "image") -> str:
    name = clean_text(name) or fallback
    name = re.sub(r"[\\/:*?\"<>|]+", "_", name)
    name = re.sub(r"\s+", "_", name).strip("._ ")
    return name or fallback


def filename_from_url(url: str, fallback: str) -> str:
    """Best-effort filename straight from a URL's own path."""
    path = urlparse(url).path
    candidate = unquote(path.rsplit("/", 1)[-1]) if path else ""
    return safe_filename(candidate, fallback)


def unique_name(base_name: str, counter: Counter) -> str:
    """Only renames on an actual collision - otherwise the given name is kept untouched."""
    base_name = safe_filename(base_name, "image")
    stem, ext = (base_name.rsplit(".", 1) + [""])[:2] if "." in base_name else (base_name, "")
    ext = f".{ext}" if ext else ""
    key = base_name.lower()
    counter[key] += 1
    return base_name if counter[key] == 1 else f"{stem}_{counter[key]}{ext}"


def has_alpha(img: Image.Image) -> bool:
    return img.mode in ("RGBA", "LA", "PA") or (img.mode == "P" and "transparency" in img.info)


def flatten_to_background(img: Image.Image, bg_color: Tuple[int, int, int] = (255, 255, 255)) -> Image.Image:
    """Paste onto a solid background wherever the image is transparent. One function
    replaces the old add_white_background()/flatten_for_jpeg() pair - same operation,
    just a configurable color."""
    if has_alpha(img):
        rgba = img.convert("RGBA")
        bg = Image.new("RGB", rgba.size, bg_color)
        bg.paste(rgba, mask=rgba.split()[3])
        return bg
    return img.convert("RGB")


def _scale_to_fit(img: Image.Image, target_w: int, target_h: int) -> Tuple[int, int]:
    """Shared 'fit inside box, keep ratio' math used by every non-cropping resize path."""
    scale = min(target_w / img.width, target_h / img.height)
    return max(1, round(img.width * scale)), max(1, round(img.height * scale))


# ---------- background / padding operations ----------

def corner_background_color(img: Image.Image) -> Tuple[int, int, int]:
    rgb = img.convert("RGB")
    w, h = rgb.size
    points = [(0, 0), (w - 1, 0), (0, h - 1), (w - 1, h - 1)]
    colors = [rgb.getpixel(p) for p in points]
    return tuple(sum(c[i] for c in colors) // 4 for i in range(3))


def remove_padding(img: Image.Image, tolerance: int = 6) -> Image.Image:
    """Trims only obvious empty space: fully transparent, or a clearly white/off-white
    border. Never resizes or intentionally cuts the product."""
    if has_alpha(img):
        rgba = img.convert("RGBA")
        alpha_bbox = rgba.split()[3].getbbox()
        return rgba.crop(alpha_bbox) if alpha_bbox else rgba

    rgb = img.convert("RGB")
    w, h = rgb.size
    step_x, step_y = max(1, w // 30), max(1, h // 30)
    border_pixels = (
        [rgb.getpixel((x, 0)) for x in range(0, w, step_x)]
        + [rgb.getpixel((x, h - 1)) for x in range(0, w, step_x)]
        + [rgb.getpixel((0, y)) for y in range(0, h, step_y)]
        + [rgb.getpixel((w - 1, y)) for y in range(0, h, step_y)]
    )
    white_like = sum(1 for r, g, b in border_pixels if r >= 245 and g >= 245 and b >= 245)
    if white_like / max(1, len(border_pixels)) < 0.85:
        return img  # border isn't clearly white - don't risk cutting the product

    white_bg = Image.new("RGB", rgb.size, (255, 255, 255))
    diff = ImageChops.difference(rgb, white_bg).convert("L")
    mask = diff.point(lambda px: 255 if px > tolerance else 0)
    bbox = mask.getbbox()
    return img.crop(bbox) if bbox else img


# ---------- resize operations ----------

def resize_fit(img: Image.Image, target_w: int, target_h: int) -> Image.Image:
    """Resize to fit inside the box, keep full image, ratio unchanged. Dimensions may
    end up smaller than target_w/target_h on one axis - that's expected, not a bug."""
    new_w, new_h = _scale_to_fit(img, target_w, target_h)
    return img.resize((new_w, new_h), Image.LANCZOS)


def resize_exact_canvas(img: Image.Image, target_w: int, target_h: int, canvas_mode: str) -> Image.Image:
    """Exact target_w x target_h, full image visible, extra space is padding."""
    new_w, new_h = _scale_to_fit(img, target_w, target_h)
    resized = img.resize((new_w, new_h), Image.LANCZOS)
    ox, oy = (target_w - new_w) // 2, (target_h - new_h) // 2

    transparent = canvas_mode == "Transparent padding"
    canvas = Image.new("RGBA" if transparent else "RGB", (target_w, target_h),
                        (255, 255, 255, 0) if transparent else (255, 255, 255))
    if resized.mode == "RGBA":
        canvas.paste(resized, (ox, oy), resized.split()[3])
    else:
        canvas.paste(resized.convert(canvas.mode), (ox, oy))
    return canvas


def resize_exact_crop(img: Image.Image, target_w: int, target_h: int) -> Image.Image:
    """Exact target_w x target_h by filling the frame and cropping overflow."""
    scale = max(target_w / img.width, target_h / img.height)
    new_w, new_h = max(1, round(img.width * scale)), max(1, round(img.height * scale))
    resized = img.resize((new_w, new_h), Image.LANCZOS)
    left, top = max(0, (new_w - target_w) // 2), max(0, (new_h - target_h) // 2)
    return resized.crop((left, top, left + target_w, top + target_h))


RESIZE_DISPATCH = {
    "By Width only": lambda img, w, h, canvas_mode: img.resize(
        (w, max(1, round(img.height * w / img.width))), Image.LANCZOS),
    "By Height only": lambda img, w, h, canvas_mode: img.resize(
        (max(1, round(img.width * h / img.height)), h), Image.LANCZOS),
    "Exact size - full image with padding": lambda img, w, h, canvas_mode: resize_exact_canvas(img, w, h, canvas_mode),
    "Exact size - fill frame / crop edges": lambda img, w, h, canvas_mode: resize_exact_crop(img, w, h),
    "Resize image only - dimensions may differ": lambda img, w, h, canvas_mode: resize_fit(img, w, h),
}


# ---------- excel / link source handling ----------

def detect_url_columns(df: pd.DataFrame) -> List[str]:
    columns = []
    for col in df.columns:
        col_name = str(col).strip().lower()
        name_hint = any(k in col_name for k in URL_KEYWORDS)
        sample = df[col].dropna().astype(str).head(40).tolist()
        url_count = sum(1 for v in sample if is_url(v))
        if url_count > 0 and (name_hint or url_count >= max(1, len(sample) // 3)):
            columns.append(col)
    return columns


def detect_filename_column(df: pd.DataFrame, url_columns: List[str]) -> Optional[str]:
    candidates = [c for c in df.columns if c not in url_columns]
    for col in candidates:
        col_name = str(col).strip().lower()
        if any(k == col_name or k in col_name for k in FILE_KEYWORDS):
            return col
    return candidates[0] if candidates else None


def build_excel_sources(df: pd.DataFrame) -> Tuple[List[Dict], List[str], Optional[str]]:
    """Filename = the value in the detected filename column, used as-is (no column
    suffix appended). If a row has multiple image columns, only the first URL per
    row keeps the bare name; extras get a short suffix so nothing is overwritten."""
    url_columns = detect_url_columns(df)
    filename_column = detect_filename_column(df, url_columns)
    sources: List[Dict] = []
    seen_urls = set()

    for row_index, row in df.iterrows():
        base_name = safe_filename(row.get(filename_column), f"row_{row_index + 1}") if filename_column else f"row_{row_index + 1}"
        row_urls = [col for col in url_columns if is_url(clean_text(row.get(col))) and clean_text(row.get(col)) not in seen_urls]
        for i, col in enumerate(row_urls):
            url = clean_text(row.get(col))
            seen_urls.add(url)
            name = base_name if i == 0 else f"{base_name}_{safe_filename(str(col), 'img')}"
            if "." not in name:
                name += ".jpg"
            sources.append({"type": "url", "url": url, "filename": name})

    return sources, [str(c) for c in url_columns], str(filename_column) if filename_column else None


def parse_pasted_links(raw_text: str) -> List[Dict]:
    """Each line: either just a URL, or 'name,url'. Blank lines ignored."""
    sources = []
    for line in raw_text.splitlines():
        line = line.strip()
        if not line:
            continue
        if "," in line:
            name_part, url_part = line.split(",", 1)
            name_part, url_part = name_part.strip(), url_part.strip()
        else:
            name_part, url_part = "", line
        if not is_url(url_part):
            continue
        filename = safe_filename(name_part) if name_part else filename_from_url(url_part, "image")
        if "." not in filename:
            filename += ".jpg"
        sources.append({"type": "url", "url": url_part, "filename": filename})
    return sources


def download_with_retry(session: requests.Session, url: str, retries: int = 3, timeout: int = 25) -> bytes:
    headers = {"User-Agent": "Mozilla/5.0 CreativeEditingAtlas/1.0"}
    last_error = None
    for attempt in range(1, retries + 1):
        try:
            response = session.get(url, timeout=timeout, headers=headers)
            response.raise_for_status()
            if not response.content:
                raise ValueError("Empty image response")
            return response.content
        except Exception as exc:
            last_error = exc
            time.sleep(0.35 * attempt)
    raise RuntimeError(str(last_error))


# ---------- main processing ----------

def process_image(
    raw_bytes: bytes,
    filename: str,
    resize_mode: str,
    target_w: int,
    target_h: int,
    output_format: Optional[str],
    quality: int,
    output_dpi: int,
    bg_mode: str,
    padding_mode: str,
    canvas_mode: str,
) -> Tuple[bytes, str, int, int, Optional[Image.Image]]:
    with Image.open(io.BytesIO(raw_bytes)) as opened:
        ImageOps.exif_transpose(opened, in_place=True)
        original_format = opened.format or "PNG"
        img = opened.copy()

    if padding_mode == "Remove padding":
        img = remove_padding(img)

    if bg_mode == "Remove background":
        if not REMBG_AVAILABLE:
            raise RuntimeError("Background removal is selected but rembg/onnxruntime is not installed.")
        removed = rembg_remove(img.convert("RGBA"), session=new_session("u2netp"))
        img = removed if isinstance(removed, Image.Image) else Image.open(io.BytesIO(removed)).convert("RGBA")
    elif bg_mode == "Add white background":
        img = flatten_to_background(img, (255, 255, 255))

    img = RESIZE_DISPATCH[resize_mode](img, target_w, target_h, canvas_mode)

    save_fmt = output_format or original_format
    if save_fmt not in ("PNG", "JPEG", "WEBP", "GIF", "BMP", "TIFF"):
        save_fmt = "PNG"
    if save_fmt == "JPEG":
        img = flatten_to_background(img, (255, 255, 255))
    elif save_fmt in ("PNG", "WEBP") and img.mode not in ("RGBA", "RGB", "L", "LA"):
        img = img.convert("RGBA" if has_alpha(img) else "RGB")

    buffer = io.BytesIO()
    save_kwargs = {
        "JPEG": {"quality": quality, "subsampling": 0, "optimize": True, "progressive": True, "dpi": (output_dpi, output_dpi)},
        "WEBP": {"quality": quality, "method": 6},
        "PNG": {"dpi": (output_dpi, output_dpi), "compress_level": 3},
    }.get(save_fmt, {})
    img.save(buffer, format=save_fmt, **save_kwargs)
    buffer.seek(0)
    output_data = buffer.read()

    stem = safe_filename(filename.rsplit(".", 1)[0] if "." in filename else filename, "image")
    output_name = f"{stem}.{EXT_MAP.get(save_fmt, save_fmt.lower())}"
    preview_img = img.copy() if img.width * img.height <= 16_000_000 else None
    width, height = img.size
    img.close()
    buffer.close()
    return output_data, output_name, width, height, preview_img


if workflow == "Resize / edit images":
    # ================= UI =================

    st.divider()
    st.subheader("Give me images")
    tab_upload, tab_excel, tab_links = st.tabs(["Upload files", "Excel with links", "Paste links"])

    with tab_upload:
        uploaded_images = st.file_uploader(
            "Upload image files", type=["png", "jpg", "jpeg", "webp", "gif", "bmp", "tiff"],
            accept_multiple_files=True,
        )
        if uploaded_images:
            st.success(f"{len(uploaded_images)} uploaded image(s) ready - original filenames will be kept.")

    with tab_excel:
        excel_file = st.file_uploader("Excel file with image links", type=["xlsx"])
        st.markdown(
            "<div class='small-note'>Image columns are detected automatically (Image1, Image2, ImageLink, URL, Photo, "
            "Front, Back, Side, etc). The filename column value is used as the output name, as-is.</div>",
            unsafe_allow_html=True,
        )

    with tab_links:
        pasted_links_raw = st.text_area(
            "One link per line. Optionally give it a name: 'my_name, https://...'",
            height=120,
            placeholder="sku123_front, https://example.com/a.jpg\nhttps://example.com/b.jpg",
        )
        st.markdown(
            "<div class='small-note'>No name given? The filename is taken from the URL itself.</div>",
            unsafe_allow_html=True,
        )

    excel_sources: List[Dict] = []
    if excel_file is not None:
        try:
            df_excel = pd.read_excel(excel_file)
            excel_sources, detected_url_columns, detected_filename_column = build_excel_sources(df_excel)
            if detected_url_columns:
                st.success(f"Found {len(excel_sources)} image link(s) from {len(detected_url_columns)} Excel column(s).")
                with st.expander("Detected Excel columns"):
                    st.write("Image columns:", ", ".join(detected_url_columns))
                    st.write("Filename column:", detected_filename_column or "Auto-generated")
            else:
                st.warning("No image URL columns were detected in the Excel file.")
        except Exception as exc:
            st.error(f"Could not read Excel file: {exc}")

    pasted_sources = parse_pasted_links(pasted_links_raw) if pasted_links_raw else []
    if pasted_sources:
        st.success(f"{len(pasted_sources)} pasted link(s) ready.")

    if not uploaded_images and not excel_sources and not pasted_sources:
        st.info("Upload images, upload an Excel file, or paste links to start.")
        st.stop()

    st.divider()
    st.subheader("Background")
    bg_mode = st.radio("Choose one", ["Keep original", "Remove background", "Add white background"], horizontal=True, index=0)
    if bg_mode == "Remove background" and not REMBG_AVAILABLE:
        st.warning("Background removal needs rembg and onnxruntime in requirements.txt.")

    st.divider()
    st.subheader("Padding")
    padding_mode = st.radio("Padding option", ["Keep padding", "Remove padding"], horizontal=True, index=0)
    st.caption("Keep padding preserves the image exactly. Remove padding only trims obvious empty white/transparent space.")

    st.divider()
    st.subheader("Dimensions")
    preset = st.selectbox("Choose marketplace recommended size", list(PRESETS.keys()))
    default_w, default_h = PRESETS[preset]
    if preset != "Custom":
        rec_w, rec_h, min_dim, max_dim, aspect_ratio = MARKETPLACE_PRESETS[preset]
        m1, m2, m3, m4 = st.columns(4)
        m1.metric("Recommended", f"{rec_w} x {rec_h}")
        m2.metric("Minimum", min_dim)
        m3.metric("Maximum", max_dim)
        m4.metric("Ratio", aspect_ratio)
    st.caption("To get exact marketplace dimensions the app uses padding or crop. Stretching is never used.")

    resize_choice = st.radio(
        "Resize mode",
        ["By Width only", "By Height only", "Exact W x H"],
        horizontal=True,
        help="By Width/Height keeps the full image ratio. Exact W x H can guarantee the selected dimensions.",
    )

    resize_mode = resize_choice
    canvas_mode = "White padding"
    if resize_choice == "Exact W x H":
        resize_mode = st.radio(
            "Exact size behaviour",
            ["Exact size - full image with padding", "Exact size - fill frame / crop edges", "Resize image only - dimensions may differ"],
            index=0,
            help="Padding keeps the full image visible. Crop fills the frame but may cut edges.",
        )
        if resize_mode == "Exact size - full image with padding":
            canvas_mode = st.radio("Padding background", ["White padding", "Transparent padding"], horizontal=True, index=0)

    col_w, col_h = st.columns(2)
    target_w, target_h = default_w, default_h
    if resize_choice in ("By Width only", "Exact W x H"):
        target_w = int(col_w.number_input("Width (px)", min_value=1, value=int(default_w), step=1))
    if resize_choice in ("By Height only", "Exact W x H"):
        target_h = int(col_h.number_input("Height (px)", min_value=1, value=int(default_h), step=1))
    output_dpi = int(st.number_input("DPI", min_value=50, max_value=1200, value=300, step=1))

    st.divider()
    st.subheader("Output Format")
    chosen_format = st.selectbox("Output format", ["Keep original format", "PNG", "JPEG", "WEBP"], index=0)
    output_format = None if chosen_format == "Keep original format" else chosen_format
    quality = 98
    if output_format in ("JPEG", "WEBP"):
        quality = st.slider(f"{output_format.title()} quality", 80, 100, 98, 1)
    else:
        st.caption("Keeping original format avoids unnecessary conversion or compression.")
    if resize_mode == "Exact size - full image with padding" and canvas_mode == "Transparent padding" and output_format == "JPEG":
        st.warning("JPEG cannot keep transparent padding. Choose PNG/WebP or use White padding.")

    st.divider()
    st.subheader("Settings")
    preview_enabled = st.checkbox("Preview first 5 processed images", value=True)
    retry_count = 3

    direct_count = len(uploaded_images) if uploaded_images else 0
    total_count = direct_count + len(excel_sources) + len(pasted_sources)
    st.info(f"Ready to process {total_count} image(s).")

    if st.button("Process & Download ZIP", type="primary", use_container_width=True):
        if total_count == 0:
            st.warning("No images found to process.")
            st.stop()

        start_time = time.time()
        progress = st.progress(0, text="Starting...")
        status_box = st.empty()
        preview_box = st.container()
        zip_buffer = io.BytesIO()
        errors, processed_rows, preview_items = [], [], []
        name_counter: Counter = Counter()
        session = requests.Session()

        def iter_sources() -> Iterable[Dict]:
            if uploaded_images:
                for file in uploaded_images:
                    yield {"type": "upload", "file": file, "filename": file.name}
            for source in excel_sources:
                yield source
            for source in pasted_sources:
                yield source

        with zipfile.ZipFile(zip_buffer, "w", compression=zipfile.ZIP_DEFLATED) as zf:
            for index, source in enumerate(iter_sources(), start=1):
                filename = source.get("filename", f"image_{index}.jpg")
                try:
                    if source["type"] == "upload":
                        status_box.info(f"Reading image {index} of {total_count}: {filename}")
                        source["file"].seek(0)
                        raw_bytes = source["file"].read()
                    else:
                        status_box.info(f"Downloading image {index} of {total_count}: {filename}")
                        raw_bytes = download_with_retry(session, source["url"], retries=retry_count)

                    status_box.info(f"Processing image {index} of {total_count}: {filename}")
                    output_data, out_name, w, h, preview_img = process_image(
                        raw_bytes=raw_bytes,
                        filename=filename,
                        resize_mode=resize_mode,
                        target_w=target_w,
                        target_h=target_h,
                        output_format=output_format,
                        quality=quality,
                        output_dpi=output_dpi,
                        bg_mode=bg_mode,
                        padding_mode=padding_mode,
                        canvas_mode=canvas_mode,
                    )
                    out_name = unique_name(out_name, name_counter)
                    zf.writestr(out_name, output_data)
                    processed_rows.append({"File": out_name, "Width": w, "Height": h, "Status": "Success"})
                    if preview_enabled and preview_img is not None and len(preview_items) < 5:
                        preview_items.append((out_name, preview_img))
                except UnidentifiedImageError:
                    errors.append({"File": filename, "Error": "File could not be opened as an image"})
                except Exception as exc:
                    errors.append({"File": filename, "Error": str(exc)})

                progress.progress(index / total_count, text=f"Processed {index} of {total_count}")
                raw_bytes = None
                if index % 25 == 0:
                    gc.collect()

            if errors:
                zf.writestr("error_report.csv", pd.DataFrame(errors).to_csv(index=False).encode("utf-8"))

        zip_buffer.seek(0)
        elapsed = time.time() - start_time
        progress.empty()
        status_box.empty()

        success_count, fail_count = len(processed_rows), len(errors)
        st.success(f"Done. {success_count} image(s) processed, {fail_count} failed.")
        c1, c2, c3 = st.columns(3)
        c1.metric("Successful", success_count)
        c2.metric("Failed", fail_count)
        c3.metric("Time", f"{elapsed:.1f}s")

        if preview_enabled and preview_items:
            with preview_box:
                st.markdown("Preview")
                for out_name, preview_img in preview_items:
                    st.image(preview_img, caption=out_name, use_container_width=True)
                    preview_img.close()

        if errors:
            with st.expander("View errors"):
                st.dataframe(pd.DataFrame(errors), use_container_width=True)
                st.caption("The ZIP includes error_report.csv.")

        st.download_button(
            label=f"Download ZIP ({success_count} images)",
            data=zip_buffer,
            file_name="creative_editing_for_atlas.zip",
            mime="application/zip",
            use_container_width=True,
            type="primary",
        )

    st.divider()
    st.caption("Default behavior keeps the original image look. No shadow is added. Cropping happens only if Remove padding is selected. Output filenames match the name you gave each image.")


else:
    # ============= PXM: download, rename, and build the Image Stack Import Template =============
    st.divider()
    st.subheader("1. Excel file")
    st.caption(
        "One row per product. Needs a Master ID column, an MPN column, and one or more image-link columns "
        "(e.g. Image1, Image2, Image3 ...)."
    )
    pxm_excel = st.file_uploader("Excel file with Master ID, MPN and image links", type=["xlsx"], key="pxm_excel")

    if pxm_excel is None:
        st.info("Upload an Excel file to continue.")
        st.stop()

    try:
        pxm_df = pd.read_excel(pxm_excel)
    except Exception as exc:
        st.error(f"Could not read Excel file: {exc}")
        st.stop()

    all_columns = [str(c) for c in pxm_df.columns]
    auto_url_cols = detect_url_columns(pxm_df)

    def _guess_col(keywords, exclude=()):
        for c in all_columns:
            lc = c.lower()
            if c not in exclude and any(k in lc for k in keywords):
                return c
        return all_columns[0] if all_columns else None

    st.divider()
    st.subheader("2. Column mapping")
    col_a, col_b = st.columns(2)
    master_id_col = col_a.selectbox(
        "Master ID column", all_columns, index=all_columns.index(_guess_col(["master", "mid"]) or all_columns[0])
    )
    mpn_col = col_b.selectbox(
        "MPN column (optional)", ["(none)"] + all_columns,
        index=(["(none)"] + all_columns).index(_guess_col(["mpn"], exclude=[master_id_col]) or "(none)"),
    )
    image_cols = st.multiselect(
        "Image link columns, in the order they should be numbered (1st = ISP_01, 2nd = ISP_02, ...)",
        [c for c in all_columns if c not in (master_id_col, mpn_col)],
        default=[c for c in auto_url_cols if c not in (master_id_col, mpn_col)],
    )
    if not image_cols:
        st.warning("Pick at least one image link column.")
        st.stop()

    st.divider()
    st.subheader("3. Marketplace & naming")
    marketplace_label = st.selectbox("Marketplace", list(PXM_MARKETPLACE_CODES.keys()) + ["Custom"])
    if marketplace_label == "Custom":
        combined_code = st.text_input("Marketplace + Region code (e.g. ALG_PL)", "").strip()
    else:
        combined_code = PXM_MARKETPLACE_CODES[marketplace_label]
        st.caption(f"Naming convention: MPN_MasterID_**{combined_code}**_ISP_XX")

    force_jpg = st.checkbox("Force convert all images to JPG", value=False)
    st.caption("PNG / WebP / GIF / BMP / TIFF will be re-encoded as JPG (transparency flattened onto white).")

    st.divider()
    st.subheader("4. Image Stack Import Template")
    import_type = st.selectbox("Import Type", ["Create/Edit", "Edit", "Remove Image", "Delete"], index=0)
    stack_template_file = st.file_uploader(
        "Optional: upload your blank Image Stack Import Template so the output matches its exact headers",
        type=["xlsx"], key="stack_template",
    )

    st.divider()
    threads = int(st.slider("Download threads", 1, 64, 12))
    save_failure_report = st.checkbox("Save failure report (.xlsx) to output", value=True)

    total_links = 0
    for _, row in pxm_df.iterrows():
        total_links += sum(1 for c in image_cols if is_url(clean_text(row.get(c))))
    st.info(f"Ready to download and rename {total_links} image(s) across {len(pxm_df)} row(s).")

    if st.button("Download, Rename & Build Template", type="primary", use_container_width=True):
        if not combined_code:
            st.warning("Enter a Marketplace + Region code first.")
            st.stop()

        jobs = []  # (master_id, mpn, url, seq)
        for _, row in pxm_df.iterrows():
            master_id = safe_filename(row.get(master_id_col), "")
            mpn = safe_filename(row.get(mpn_col), "") if mpn_col != "(none)" else ""
            seq = 0
            for col in image_cols:
                url = clean_text(row.get(col))
                if not is_url(url):
                    continue
                seq += 1
                jobs.append((master_id, mpn, url, seq))

        progress = st.progress(0, text="Starting...")
        status_box = st.empty()
        zip_buffer = io.BytesIO()
        stack_rows, errors = [], []
        session = requests.Session()

        def _download_one(job):
            master_id, mpn, url, seq = job
            name_prefix = "_".join(p for p in (mpn, master_id) if p)
            ext_guess = url.rsplit(".", 1)[-1].lower() if "." in url.rsplit("/", 1)[-1] else "jpg"
            ext_guess = ext_guess if ext_guess in ("jpg", "jpeg", "png", "webp", "gif", "bmp", "tif", "tiff") else "jpg"
            out_ext = "jpg" if force_jpg else ext_guess
            filename = f"{name_prefix}_{combined_code}_ISP_{seq:02d}.{out_ext}"
            try:
                raw = download_with_retry(session, url, retries=3)
                if force_jpg:
                    with Image.open(io.BytesIO(raw)) as im:
                        ImageOps.exif_transpose(im, in_place=True)
                        im = flatten_to_background(im, (255, 255, 255))
                        buf = io.BytesIO()
                        im.save(buf, format="JPEG", quality=95, optimize=True)
                        raw = buf.getvalue()
                return {"ok": True, "master_id": master_id, "filename": filename, "seq": seq, "data": raw}
            except Exception as exc:
                return {"ok": False, "master_id": master_id, "filename": filename, "seq": seq, "error": str(exc)}

        from concurrent.futures import ThreadPoolExecutor, as_completed

        results = []
        with ThreadPoolExecutor(max_workers=max(1, threads)) as executor:
            futures = {executor.submit(_download_one, job): job for job in jobs}
            done = 0
            for future in as_completed(futures):
                result = future.result()
                results.append(result)
                done += 1
                status_box.info(f"Downloaded {done} of {len(jobs)}: {result['filename']}")
                progress.progress(done / max(1, len(jobs)), text=f"{done} of {len(jobs)}")

        results.sort(key=lambda r: (r["master_id"], r["seq"]))
        name_counter: Counter = Counter()
        with zipfile.ZipFile(zip_buffer, "w", compression=zipfile.ZIP_DEFLATED) as zf:
            for r in results:
                if r["ok"]:
                    final_name = unique_name(r["filename"], name_counter)
                    zf.writestr(final_name, r["data"])
                    stack_rows.append([import_type, r["master_id"], combined_code, final_name, r["seq"]])
                else:
                    errors.append({"Master ID": r["master_id"], "Filename": r["filename"], "Error": r["error"]})
            if errors and save_failure_report:
                zf.writestr("failure_report.csv", pd.DataFrame(errors).to_csv(index=False).encode("utf-8"))

        zip_buffer.seek(0)
        progress.empty()
        status_box.empty()

        # Build the Image Stack Import Template workbook
        import openpyxl

        if stack_template_file is not None:
            stack_template_file.seek(0)
            wb = openpyxl.load_workbook(stack_template_file)
            sheet_name = next((n for n in wb.sheetnames if "template" in n.lower()), wb.sheetnames[-1])
            ws = wb[sheet_name]
            header_row_idx = None
            for i, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), start=1):
                if row and str(row[0]).strip().lower() == "import type":
                    header_row_idx = i
                    break
            start_row = (header_row_idx or 1) + 1
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Image Stack Import Template"
            ws.append(STACK_TEMPLATE_HEADERS)
            start_row = 2

        for offset, stack_row in enumerate(stack_rows):
            for col_idx, value in enumerate(stack_row, start=1):
                ws.cell(row=start_row + offset, column=col_idx, value=value)

        template_buffer = io.BytesIO()
        wb.save(template_buffer)
        template_buffer.seek(0)

        success_count, fail_count = len(stack_rows), len(errors)
        st.success(f"Done. {success_count} image(s) downloaded and renamed, {fail_count} failed.")
        c1, c2 = st.columns(2)
        c1.metric("Successful", success_count)
        c2.metric("Failed", fail_count)

        if errors:
            with st.expander("View errors"):
                st.dataframe(pd.DataFrame(errors), use_container_width=True)

        dl1, dl2 = st.columns(2)
        with dl1:
            st.download_button(
                "Download renamed images (ZIP)", data=zip_buffer,
                file_name="pxm_renamed_images.zip", mime="application/zip", use_container_width=True,
            )
        with dl2:
            st.download_button(
                "Download Image Stack Template (.xlsx)", data=template_buffer,
                file_name="image_stack_import_template.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

    st.divider()
    st.caption(
        "Filename = MPN_MasterID_MarketplaceRegionCode_ISP_XX. Image Stack Group in the template always "
        "matches the Marketplace + Region code you picked above, so the two outputs stay in sync."
    )
